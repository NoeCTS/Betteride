#!/usr/bin/env python3

from __future__ import annotations

import csv
import io
import json
import shutil
import sqlite3
import struct
import urllib.error
import urllib.request
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
NORMALIZED_DIR = ROOT / "data" / "normalized"

DOWNLOADS = {
    "berlin_bike_counters_raw.xlsx": "https://www.berlin.de/sen/uvk/_assets/verkehr/verkehrsplanung/radverkehr/weitere-radinfrastruktur/zaehlstellen-und-fahrradbarometer/gesamtdatei-stundenwerte.xlsx?ts=1737968619",
    "berlin_osm_bicycle_shops_source.gpkg.zip": "https://download.geofabrik.de/europe/germany/berlin-latest-free.gpkg.zip",
    "vbb_station_access_source.zip": "https://www.vbb.de/fileadmin/user_upload/VBB/Dokumente/API-Datensaetze/UMBW.zip",
    "vbb_gtfs.zip": "https://www.vbb.de/vbbgtfs",
}

GOOGLE_PLACES_HEADERS = [
    "shop_name",
    "google_place_id",
    "rating",
    "user_ratings_total",
    "formatted_address",
    "website",
    "phone",
    "lat",
    "lon",
]

YELP_HEADERS = [
    "shop_name",
    "yelp_id",
    "rating",
    "review_count",
    "address",
    "phone",
    "website",
    "lat",
    "lon",
]

FOOTFALL_HEADERS = [
    "station_id",
    "station_name",
    "entry_exit_count",
    "period",
    "source_url",
    "status",
    "note",
]


def main() -> None:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    NORMALIZED_DIR.mkdir(parents=True, exist_ok=True)

    for filename, url in DOWNLOADS.items():
        ensure_download(RAW_DIR / filename, url)

    ensure_gpkg_extracted()

    bike_counter_summary = build_bike_counters()
    repair_shop_summary = build_bike_repair_shops()
    station_access_summary = build_station_access()
    gtfs_summary = build_vbb_stops()
    placeholder_summary = build_placeholders()

    summary = {
        "generatedAt": iso_now(),
        "datasets": [
            {
                "id": "berlin-bike-counters",
                "label": "Berlin bike counters",
                "status": "ready",
                "countLabel": f"{bike_counter_summary['location_count']:,} locations / {bike_counter_summary['hourly_rows']:,} hourly rows",
                "detail": f"Official workbook converted from {bike_counter_summary['year_start']} to {bike_counter_summary['year_end']}.",
            },
            {
                "id": "berlin-bike-repair-shops",
                "label": "Berlin bike repair shops",
                "status": "ready",
                "countLabel": f"{repair_shop_summary['feature_count']:,} shops",
                "detail": "Geofabrik Berlin GPKG filtered on the bicycle shop class and normalized into CSV and GeoJSON for supply mapping.",
            },
            {
                "id": "vbb-station-access",
                "label": "VBB station access points",
                "status": "ready",
                "countLabel": f"{station_access_summary['feature_count']:,} access points",
                "detail": "Official UMBW export normalized from semicolon-delimited legacy CSV into UTF-8 CSV and GeoJSON.",
            },
            {
                "id": "vbb-gtfs-stops",
                "label": "VBB GTFS stops",
                "status": "ready",
                "countLabel": f"{gtfs_summary['feature_count']:,} stops",
                "detail": "Weekly GTFS static stops exported to CSV and GeoJSON for stop/station joins.",
            },
            {
                "id": "google-places",
                "label": "Google Places ratings",
                "status": "placeholder",
                "countLabel": "header-only CSV",
                "detail": "API key and billing are still required before ratings and review counts can be fetched.",
            },
            {
                "id": "yelp-ratings",
                "label": "Yelp ratings",
                "status": "placeholder",
                "countLabel": "header-only CSV",
                "detail": "Developer credentials are still required before Yelp review metadata can be fetched.",
            },
            {
                "id": "station-footfall",
                "label": "Station footfall",
                "status": "placeholder",
                "countLabel": "placeholder CSV",
                "detail": "No verified public bulk passenger-count file is staged yet; corridor strength still needs a proxy or licensed feed.",
            },
        ],
        "files": [
            "data/raw/berlin_bike_counters_raw.xlsx",
            "data/raw/berlin_bike_counters_hourly.csv",
            "data/raw/berlin_bike_counters_locations.geojson",
            "data/raw/berlin_osm_bicycle_shops_source.gpkg",
            "data/raw/berlin_bike_repair_shops.geojson",
            "data/raw/berlin_bike_repair_shops.csv",
            "data/raw/vbb_station_access_source.zip",
            "data/raw/vbb_station_access_points.geojson",
            "data/raw/vbb_station_access_points.csv",
            "data/raw/vbb_gtfs.zip",
            "data/raw/vbb_stops.csv",
            "data/raw/vbb_stops.geojson",
            "data/raw/berlin_bike_shops_google_places.csv",
            "data/raw/berlin_bike_shops_yelp.csv",
            "data/raw/berlin_station_footfall.csv",
        ],
        "notes": [
            "The dashboard map still uses lightweight seed layers for scoring and rendering stability.",
            "Raw market files now live under data/raw and can be promoted into a heavier-weight scoring pipeline later.",
            "Google Places, Yelp, and station footfall remain placeholders until credentials or a verified bulk source are available.",
        ],
    }

    write_json(
        NORMALIZED_DIR / "raw-market-data-summary.json",
        summary,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def ensure_download(destination: Path, url: str) -> None:
    if destination.exists():
        print(f"Using existing {destination.relative_to(ROOT)}")
        return

    print(f"Downloading {url} -> {destination.relative_to(ROOT)}")
    request = urllib.request.Request(url, headers={"User-Agent": "betteride-ground-signal/1.0"})

    try:
        with urllib.request.urlopen(request) as response, destination.open("wb") as output:
            shutil.copyfileobj(response, output)
    except urllib.error.URLError as error:
        raise SystemExit(f"Failed to download {url}: {error}") from error


def ensure_gpkg_extracted() -> Path:
    extracted_path = RAW_DIR / "berlin_osm_bicycle_shops_source.gpkg"
    legacy_path = RAW_DIR / "berlin.gpkg"

    if extracted_path.exists():
        return extracted_path

    if legacy_path.exists():
        legacy_path.rename(extracted_path)
        return extracted_path

    archive_path = RAW_DIR / "berlin_osm_bicycle_shops_source.gpkg.zip"
    with zipfile.ZipFile(archive_path) as archive:
        with archive.open("berlin.gpkg") as source, extracted_path.open("wb") as target:
            shutil.copyfileobj(source, target)

    return extracted_path


def build_bike_counters() -> dict[str, int]:
    workbook = load_workbook(
        RAW_DIR / "berlin_bike_counters_raw.xlsx",
        read_only=True,
        data_only=True,
    )

    locations_sheet = workbook["Standortdaten"]
    location_features = []

    rows = locations_sheet.iter_rows(values_only=True)
    next(rows, None)

    for row in rows:
        counter_id = clean_string(row[0])
        if not counter_id:
            continue

        name = clean_string(row[1])
        lat = to_float(row[2])
        lon = to_float(row[3])
        install_date = to_iso_date(row[4])
        if lat is None or lon is None:
            continue

        location_features.append(
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [lon, lat]},
                "properties": {
                    "counter_id": counter_id,
                    "name": name,
                    "install_date": install_date,
                    "lat": lat,
                    "lon": lon,
                },
            }
        )

    write_geojson(RAW_DIR / "berlin_bike_counters_locations.geojson", location_features)

    year_sheets = [name for name in workbook.sheetnames if name.startswith("Jahresdatei ")]
    counter_columns: list[str] = []
    counter_metadata: dict[str, dict[str, str]] = {}

    for sheet_name in year_sheets:
        sheet = workbook[sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        for value in header_row[1:]:
            counter_id, install_date = parse_counter_header(value)
            if not counter_id:
                continue
            if counter_id not in counter_metadata:
                counter_metadata[counter_id] = {"install_date": install_date}
                counter_columns.append(counter_id)

    hourly_rows = 0
    output_path = RAW_DIR / "berlin_bike_counters_hourly.csv"
    with output_path.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(["timestamp", *counter_columns])

        for sheet_name in year_sheets:
            sheet = workbook[sheet_name]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            sheet_counter_ids = [parse_counter_header(value)[0] for value in header_row[1:]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                timestamp = to_iso_timestamp(row[0])
                if not timestamp:
                    continue

                row_values = {counter_id: value for counter_id, value in zip(sheet_counter_ids, row[1:]) if counter_id}
                writer.writerow(
                    [
                        timestamp,
                        *[
                            "" if row_values.get(counter_id) is None else row_values[counter_id]
                            for counter_id in counter_columns
                        ],
                    ]
                )
                hourly_rows += 1

    years = sorted(int(name.split()[-1]) for name in year_sheets)
    return {
        "location_count": len(location_features),
        "hourly_rows": hourly_rows,
        "year_start": years[0],
        "year_end": years[-1],
    }


def build_bike_repair_shops() -> dict[str, int]:
    gpkg_path = ensure_gpkg_extracted()
    records: list[dict[str, object]] = []

    with sqlite3.connect(gpkg_path) as connection:
        for table_name in ("gis_osm_pois_free", "gis_osm_pois_a_free"):
            query = f"""
                SELECT osm_id, name, geom
                FROM {table_name}
                WHERE fclass = 'bicycle_shop'
            """

            for osm_id, name, geom in connection.execute(query):
                lat, lon = extract_gpkg_coordinates(geom)
                if lat is None or lon is None:
                    continue

                records.append(
                    {
                        "shop_name": clean_string(name),
                        "lat": lat,
                        "lon": lon,
                        "opening_hours": "",
                        "website": "",
                        "phone": "",
                        "shop": "bicycle",
                        "craft": "",
                        "osm_id": clean_string(osm_id),
                    }
                )

    records = dedupe_records(records)

    records.sort(key=lambda item: (item["shop_name"] or "", item["osm_id"]))

    with (RAW_DIR / "berlin_bike_repair_shops.csv").open("w", encoding="utf-8", newline="") as output:
        writer = csv.DictWriter(
            output,
            fieldnames=[
                "shop_name",
                "lat",
                "lon",
                "opening_hours",
                "website",
                "phone",
                "shop",
                "craft",
                "osm_id",
            ],
        )
        writer.writeheader()
        writer.writerows(records)

    features = [
        {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [record["lon"], record["lat"]]},
            "properties": record,
        }
        for record in records
    ]
    write_geojson(RAW_DIR / "berlin_bike_repair_shops.geojson", features)

    return {"feature_count": len(records)}


def build_station_access() -> dict[str, int]:
    fields = [
        "station_name",
        "access_name",
        "access_type",
        "access_point_number",
        "level",
        "lon",
        "lat",
        "station_reference_id",
        "access_point_id",
    ]

    csv_path = RAW_DIR / "vbb_station_access_points.csv"
    geojson_path = RAW_DIR / "vbb_station_access_points.geojson"

    with zipfile.ZipFile(RAW_DIR / "vbb_station_access_source.zip") as archive:
        with archive.open("UMBW.CSV") as source:
            text_stream = io.TextIOWrapper(source, encoding="cp1252", newline="")
            reader = csv.DictReader(text_stream, delimiter=";")

            with csv_path.open("w", encoding="utf-8", newline="") as csv_output:
                writer = csv.DictWriter(csv_output, fieldnames=fields)
                writer.writeheader()

                feature_count = 0
                first_feature = True
                with geojson_path.open("w", encoding="utf-8") as geojson_output:
                    geojson_output.write('{"type":"FeatureCollection","features":[')

                    for row in reader:
                        lat = parse_decimal_comma(row.get("Y-Koordinate"))
                        lon = parse_decimal_comma(row.get("X-Koordinate"))
                        if lat is None or lon is None:
                            continue

                        record = {
                            "station_name": clean_string(row.get("Bauwerk Name")),
                            "access_name": clean_string(row.get("Bauwerkselement Name")),
                            "access_type": clean_string(row.get("Bauwerkselement Typ")),
                            "access_point_number": clean_string(row.get("Bauwerkselement Nummer")),
                            "level": clean_string(row.get("Bauwerkselement Niveau")),
                            "lon": lon,
                            "lat": lat,
                            "station_reference_id": clean_string(row.get("Bauwerksreferenzort Nummer")),
                            "access_point_id": clean_string(row.get("Bauwerkselement Exportnummer")),
                        }

                        writer.writerow(record)
                        feature = {
                            "type": "Feature",
                            "geometry": {"type": "Point", "coordinates": [lon, lat]},
                            "properties": record,
                        }

                        if not first_feature:
                            geojson_output.write(",")
                        json.dump(feature, geojson_output, ensure_ascii=False)
                        first_feature = False
                        feature_count += 1

                    geojson_output.write("]}\n")

    return {"feature_count": feature_count}


def build_vbb_stops() -> dict[str, int]:
    source_zip = RAW_DIR / "vbb_gtfs.zip"
    output_csv = RAW_DIR / "vbb_stops.csv"
    output_geojson = RAW_DIR / "vbb_stops.geojson"

    with zipfile.ZipFile(source_zip) as archive:
        with archive.open("stops.txt") as source, output_csv.open("wb") as target:
            shutil.copyfileobj(source, target)

        with archive.open("stops.txt") as source:
            text_stream = io.TextIOWrapper(source, encoding="utf-8-sig", newline="")
            reader = csv.DictReader(text_stream)

            feature_count = 0
            first_feature = True
            with output_geojson.open("w", encoding="utf-8") as geojson_output:
                geojson_output.write('{"type":"FeatureCollection","features":[')

                for row in reader:
                    lat = to_float(row.get("stop_lat"))
                    lon = to_float(row.get("stop_lon"))
                    if lat is None or lon is None:
                        continue

                    properties = {key: clean_string(value) for key, value in row.items()}
                    feature = {
                        "type": "Feature",
                        "geometry": {"type": "Point", "coordinates": [lon, lat]},
                        "properties": properties,
                    }

                    if not first_feature:
                        geojson_output.write(",")
                    json.dump(feature, geojson_output, ensure_ascii=False)
                    first_feature = False
                    feature_count += 1

                geojson_output.write("]}\n")

    return {"feature_count": feature_count}


def build_placeholders() -> dict[str, str]:
    write_csv_headers(
        RAW_DIR / "berlin_bike_shops_google_places.csv",
        GOOGLE_PLACES_HEADERS,
    )
    write_csv_headers(
        RAW_DIR / "berlin_bike_shops_yelp.csv",
        YELP_HEADERS,
    )
    write_placeholder_footfall(
        RAW_DIR / "berlin_station_footfall.csv",
    )
    return {"status": "ok"}


def write_csv_headers(path: Path, headers: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(headers)


def write_placeholder_footfall(path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(FOOTFALL_HEADERS)
        writer.writerow(
            [
                "",
                "",
                "",
                "",
                "https://unternehmen.vbb.de/en/digital-services/datasets/",
                "not_verified_public_bulk_source",
                "Placeholder row only. Use GTFS stop centrality and station access geometry until a verified public or licensed passenger-count feed is available.",
            ]
        )


def write_geojson(path: Path, features: list[dict]) -> None:
    with path.open("w", encoding="utf-8") as output:
        json.dump({"type": "FeatureCollection", "features": features}, output, ensure_ascii=False)
        output.write("\n")


def write_json(path: Path, payload: dict) -> None:
    with path.open("w", encoding="utf-8") as output:
        json.dump(payload, output, ensure_ascii=False, indent=2)
        output.write("\n")


def parse_counter_header(value: object) -> tuple[str, str]:
    text = clean_string(value)
    if not text or text.startswith("Zählstelle"):
        return "", ""

    parts = [part.strip() for part in text.splitlines() if part and part.strip()]
    counter_id = parts[0] if parts else ""
    install_date = parts[1] if len(parts) > 1 else ""
    return counter_id, install_date


def parse_decimal_comma(value: object) -> float | None:
    if value is None:
        return None
    return to_float(str(value).replace(".", "").replace(",", ".") if str(value).count(",") == 1 and str(value).count(".") > 1 else str(value).replace(",", "."))


def to_float(value: object) -> float | None:
    if value in (None, ""):
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_iso_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return clean_string(value)


def to_iso_timestamp(value: object) -> str:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat(timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat(timespec="seconds")
    return clean_string(value)


def clean_string(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def iso_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def dedupe_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    unique_records: dict[str, dict[str, object]] = {}
    for record in records:
        unique_records[str(record["osm_id"])] = record
    return list(unique_records.values())


def extract_gpkg_coordinates(blob: bytes | None) -> tuple[float | None, float | None]:
    if not blob or blob[:2] != b"GP":
        return None, None

    flags = blob[3]
    header_endianness = "<" if flags & 1 else ">"
    envelope_indicator = (flags >> 1) & 0b111
    envelope_bytes = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}.get(envelope_indicator, 0)
    wkb_offset = 8 + envelope_bytes

    if len(blob) < wkb_offset + 5:
        return None, None

    wkb_endianness = "<" if blob[wkb_offset] == 1 else ">"
    geom_type = struct.unpack(f"{wkb_endianness}I", blob[wkb_offset + 1 : wkb_offset + 5])[0]

    if geom_type == 1 and len(blob) >= wkb_offset + 21:
        lon, lat = struct.unpack(f"{wkb_endianness}2d", blob[wkb_offset + 5 : wkb_offset + 21])
        return lat, lon

    if envelope_bytes >= 32:
        min_x, max_x, min_y, max_y = struct.unpack(f"{header_endianness}4d", blob[8:40])
        return (min_y + max_y) / 2, (min_x + max_x) / 2

    return None, None


if __name__ == "__main__":
    main()
